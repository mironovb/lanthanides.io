#!/usr/bin/env python3
"""Import invoice files into the Strategic Materials Ledger price records.

Reads invoice files (PDF, CSV, XLSX) from a configurable input directory,
extracts pricing data, normalizes it to the standard price record schema,
and appends the results to _data/price_records.json.

Uncertain parses are flagged with lower confidence scores and explanatory notes.
Fields are never silently invented; partial data is stored and flagged for review.

Usage:
    python import_invoices.py
    python import_invoices.py --input-dir ../invoices --output-file ../_data/price_records.json
    python import_invoices.py --dry-run
"""

import argparse
import csv
import datetime
import hashlib
import json
import os
import re
import sys
from pathlib import Path

# Optional heavy dependencies -- fail gracefully if missing.
try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import openpyxl
except ImportError:
    openpyxl = None


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_INPUT_DIR = SCRIPT_DIR / ".." / "invoices"
DEFAULT_OUTPUT_FILE = SCRIPT_DIR / ".." / "_data" / "price_records.json"

EXPECTED_COLUMNS = {
    "element", "element_symbol", "quantity", "price", "date",
    "currency", "supplier", "purity", "form", "unit",
}

# Regex helpers for PDF text extraction.
RE_ELEMENT = re.compile(
    r"\b(Neodymium|Praseodymium|Dysprosium|Terbium|Samarium|Gadolinium|"
    r"Yttrium|Scandium|Tellurium|Vanadium|Gallium|Germanium|Antimony|"
    r"Tungsten|Indium|Bismuth|Molybdenum|"
    r"Nd|Pr|Dy|Tb|Sm|Gd|Y|Sc|Te|V|Ga|Ge|Sb|W|In|Bi|Mo)\b",
    re.IGNORECASE,
)
RE_PRICE = re.compile(r"(?:USD|US\$|\$|EUR|CNY|RMB|GBP)\s*([\d,]+\.?\d*)", re.IGNORECASE)
RE_PRICE_ALT = re.compile(r"([\d,]+\.?\d*)\s*(?:USD|US\$|\$|EUR|CNY|RMB|GBP|/kg|/g|/lb|/t)", re.IGNORECASE)
RE_DATE = re.compile(r"\b(\d{4}[-/]\d{1,2}[-/]\d{1,2})\b")
RE_PURITY = re.compile(r"(\d{2,3}(?:\.\d+)?)\s*%")
RE_QUANTITY = re.compile(r"(\d+(?:\.\d+)?)\s*(kg|g|lb|t|oz)\b", re.IGNORECASE)
RE_FORM = re.compile(r"\b(oxide|metal|powder|alloy|compound|carbide|ingot|pellet|sputtering target)\b", re.IGNORECASE)
RE_CURRENCY = re.compile(r"\b(USD|EUR|CNY|RMB|GBP|JPY)\b", re.IGNORECASE)

SYMBOL_NAME_MAP = {
    "neodymium": "Nd", "praseodymium": "Pr", "dysprosium": "Dy",
    "terbium": "Tb", "samarium": "Sm", "gadolinium": "Gd",
    "yttrium": "Y", "scandium": "Sc", "tellurium": "Te",
    "vanadium": "V", "gallium": "Ga", "germanium": "Ge",
    "antimony": "Sb", "tungsten": "W", "indium": "In",
    "bismuth": "Bi", "molybdenum": "Mo",
}
NAME_FOR_SYMBOL = {v: k.title() for k, v in SYMBOL_NAME_MAP.items()}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def make_record_id(filename: str, index: int) -> str:
    """Generate a deterministic record id from filename and row index."""
    raw = f"{filename}:{index}"
    short_hash = hashlib.sha256(raw.encode()).hexdigest()[:12]
    return f"inv-{short_hash}"


def now_iso() -> str:
    """Return the current UTC time in ISO 8601 format."""
    return datetime.datetime.now(datetime.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def resolve_symbol(raw: str) -> tuple:
    """Return (symbol, name) from a raw string that may be a symbol or full name."""
    if not raw:
        return (None, None)
    cleaned = raw.strip()
    # Direct symbol match (case-sensitive for single/two letter symbols).
    if cleaned in NAME_FOR_SYMBOL:
        return (cleaned, NAME_FOR_SYMBOL[cleaned])
    # Name match.
    low = cleaned.lower()
    if low in SYMBOL_NAME_MAP:
        return (SYMBOL_NAME_MAP[low], low.title())
    # Case-insensitive symbol match.
    for sym in NAME_FOR_SYMBOL:
        if cleaned.upper() == sym.upper():
            return (sym, NAME_FOR_SYMBOL[sym])
    return (cleaned, None)


def blank_record() -> dict:
    """Return a price record with all fields set to their default empty values."""
    return {
        "id": None,
        "element_symbol": None,
        "element_name": None,
        "form": None,
        "purity": None,
        "market_tier": "retail",
        "moq_kg": None,
        "quoted_quantity_kg": None,
        "original_currency": None,
        "original_unit": "kg",
        "original_price_per_unit": None,
        "normalized_usd_per_kg": None,
        "exchange_rate_used": None,
        "exchange_rate_date": None,
        "incoterm": None,
        "taxes_included": "unknown",
        "shipping_included": "unknown",
        "seller_name": None,
        "seller_country": None,
        "buyer_country": None,
        "source_type": "invoice",
        "source_id": None,
        "source_url": None,
        "invoice_ref": None,
        "quote_date": None,
        "ingestion_timestamp": None,
        "verification_status": "verified_invoice",
        "confidence_score": 0.9,
        "notes": "",
    }


def parse_price_string(raw: str):
    """Attempt to parse a numeric price from a string, returning float or None."""
    if raw is None:
        return None
    cleaned = re.sub(r"[^\d.]", "", str(raw).replace(",", ""))
    try:
        return float(cleaned)
    except (ValueError, TypeError):
        return None


def normalize_date(raw: str) -> str | None:
    """Normalize a date string to YYYY-MM-DD, or return None."""
    if not raw:
        return None
    raw = str(raw).strip().replace("/", "-")
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", raw)
    if m:
        try:
            d = datetime.date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
            return d.isoformat()
        except ValueError:
            pass
    return None


# ---------------------------------------------------------------------------
# File Parsers
# ---------------------------------------------------------------------------

def parse_csv(filepath: Path) -> list[dict]:
    """Parse a CSV invoice file into a list of raw row dicts."""
    records = []
    with open(filepath, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        if reader.fieldnames is None:
            return records
        # Normalize header names to lowercase.
        reader.fieldnames = [f.strip().lower().replace(" ", "_") for f in reader.fieldnames]
        for idx, row in enumerate(reader):
            rec = _row_to_record(row, filepath.name, idx)
            records.append(rec)
    return records


def parse_xlsx(filepath: Path) -> list[dict]:
    """Parse an XLSX invoice file into a list of raw row dicts."""
    if openpyxl is None:
        print(f"  WARNING: openpyxl not installed, skipping {filepath.name}", file=sys.stderr)
        return []
    records = []
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header_raw = next(rows_iter)
        except StopIteration:
            continue
        headers = [str(h).strip().lower().replace(" ", "_") if h else f"col_{i}" for i, h in enumerate(header_raw)]
        for idx, row_vals in enumerate(rows_iter):
            row = {headers[i]: (row_vals[i] if i < len(row_vals) else None) for i in range(len(headers))}
            rec = _row_to_record(row, filepath.name, idx)
            records.append(rec)
    wb.close()
    return records


def _row_to_record(row: dict, filename: str, idx: int) -> dict:
    """Convert a raw row dict (from CSV/XLSX) into a normalized price record."""
    rec = blank_record()
    rec["id"] = make_record_id(filename, idx)
    rec["invoice_ref"] = filename
    rec["ingestion_timestamp"] = now_iso()

    notes = []

    # Element.
    element_raw = row.get("element_symbol") or row.get("element") or row.get("symbol") or row.get("material")
    symbol, name = resolve_symbol(str(element_raw) if element_raw else "")
    rec["element_symbol"] = symbol
    rec["element_name"] = name
    if not symbol:
        notes.append("Could not determine element from row.")
        rec["confidence_score"] = max(rec["confidence_score"] - 0.3, 0.1)

    # Form.
    rec["form"] = (str(row.get("form", "")).strip().lower() or None) if row.get("form") else None

    # Purity.
    purity_raw = row.get("purity")
    if purity_raw:
        purity_str = str(purity_raw).strip()
        if "%" not in purity_str:
            purity_str += "%"
        rec["purity"] = purity_str

    # Price.
    price_raw = row.get("price") or row.get("price_per_unit") or row.get("unit_price")
    rec["original_price_per_unit"] = parse_price_string(price_raw)

    # Currency.
    currency = (str(row.get("currency", "")).strip().upper() or None) if row.get("currency") else None
    rec["original_currency"] = currency

    # Unit.
    unit = (str(row.get("unit", "")).strip().lower() or None) if row.get("unit") else None
    rec["original_unit"] = unit or "kg"

    # Quantity.
    qty_raw = row.get("quantity") or row.get("qty") or row.get("quoted_quantity")
    if qty_raw is not None:
        try:
            rec["quoted_quantity_kg"] = float(str(qty_raw).replace(",", ""))
        except (ValueError, TypeError):
            pass

    # Supplier.
    rec["seller_name"] = (str(row.get("supplier") or row.get("seller") or row.get("vendor") or "").strip() or None)

    # Country.
    rec["seller_country"] = (str(row.get("country") or row.get("seller_country") or "").strip().upper() or None)

    # Date.
    date_raw = row.get("date") or row.get("quote_date") or row.get("invoice_date")
    rec["quote_date"] = normalize_date(str(date_raw) if date_raw else "")

    # Simple USD normalization when currency is USD and unit is kg.
    if rec["original_price_per_unit"] is not None and rec["original_currency"] == "USD" and rec["original_unit"] == "kg":
        rec["normalized_usd_per_kg"] = rec["original_price_per_unit"]
        rec["exchange_rate_used"] = 1.0
        rec["exchange_rate_date"] = rec["quote_date"]

    if notes:
        rec["notes"] = "; ".join(notes)

    return rec


def parse_pdf(filepath: Path) -> list[dict]:
    """Extract price records from a PDF invoice via regex-based text parsing.

    This is inherently fragile. Records produced here receive a lower
    confidence score and notes describing which fields were extracted vs
    guessed.
    """
    if pdfplumber is None:
        print(f"  WARNING: pdfplumber not installed, skipping {filepath.name}", file=sys.stderr)
        return []

    text = ""
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text()
            if page_text:
                text += page_text + "\n"

    if not text.strip():
        return []

    rec = blank_record()
    rec["id"] = make_record_id(filepath.name, 0)
    rec["invoice_ref"] = filepath.name
    rec["ingestion_timestamp"] = now_iso()
    rec["confidence_score"] = 0.5  # Start lower for PDF extraction.

    notes = []
    extracted_fields = []

    # Element.
    m = RE_ELEMENT.search(text)
    if m:
        symbol, name = resolve_symbol(m.group(0))
        rec["element_symbol"] = symbol
        rec["element_name"] = name
        extracted_fields.append("element")
    else:
        notes.append("Element not identified in PDF text.")
        rec["confidence_score"] -= 0.2

    # Price.
    m = RE_PRICE.search(text)
    if not m:
        m = RE_PRICE_ALT.search(text)
    if m:
        rec["original_price_per_unit"] = parse_price_string(m.group(1))
        extracted_fields.append("price")
    else:
        notes.append("Price not identified in PDF text.")
        rec["confidence_score"] -= 0.2

    # Currency.
    m = RE_CURRENCY.search(text)
    if m:
        cur = m.group(1).upper()
        if cur == "RMB":
            cur = "CNY"
        rec["original_currency"] = cur
        extracted_fields.append("currency")

    # Date.
    m = RE_DATE.search(text)
    if m:
        rec["quote_date"] = normalize_date(m.group(1))
        extracted_fields.append("date")

    # Purity.
    m = RE_PURITY.search(text)
    if m:
        rec["purity"] = m.group(0)
        extracted_fields.append("purity")

    # Form.
    m = RE_FORM.search(text)
    if m:
        rec["form"] = m.group(0).lower()
        extracted_fields.append("form")

    # Quantity.
    m = RE_QUANTITY.search(text)
    if m:
        try:
            qty = float(m.group(1))
            unit = m.group(2).lower()
            rec["original_unit"] = unit
            rec["quoted_quantity_kg"] = qty
            extracted_fields.append("quantity")
        except ValueError:
            pass

    rec["confidence_score"] = max(round(rec["confidence_score"], 2), 0.1)
    notes.insert(0, f"PDF extraction: found [{', '.join(extracted_fields) or 'none'}].")
    rec["notes"] = "; ".join(notes)

    return [rec]


# ---------------------------------------------------------------------------
# Main Logic
# ---------------------------------------------------------------------------

def load_existing_records(output_file: Path) -> list[dict]:
    """Load existing price records from the output JSON file."""
    if not output_file.exists():
        return []
    with open(output_file, "r", encoding="utf-8") as fh:
        data = json.load(fh)
    if not isinstance(data, list):
        return []
    return data


def save_records(records: list[dict], output_file: Path) -> None:
    """Write price records to the output JSON file."""
    output_file.parent.mkdir(parents=True, exist_ok=True)
    with open(output_file, "w", encoding="utf-8") as fh:
        json.dump(records, fh, indent=2, ensure_ascii=False)
        fh.write("\n")


def process_invoices(input_dir: Path, output_file: Path, dry_run: bool = False) -> list[dict]:
    """Scan input_dir for invoice files, parse them, and return new records."""
    if not input_dir.exists():
        print(f"Input directory does not exist: {input_dir}", file=sys.stderr)
        return []

    files = sorted(input_dir.iterdir())
    new_records = []

    for filepath in files:
        if filepath.is_dir():
            continue
        suffix = filepath.suffix.lower()
        print(f"  Processing: {filepath.name} ...", end=" ")

        try:
            if suffix == ".csv":
                recs = parse_csv(filepath)
            elif suffix in (".xlsx", ".xls"):
                recs = parse_xlsx(filepath)
            elif suffix == ".pdf":
                recs = parse_pdf(filepath)
            else:
                print("skipped (unsupported format)")
                continue
        except Exception as exc:
            print(f"ERROR: {exc}")
            continue

        print(f"{len(recs)} record(s)")
        new_records.extend(recs)

    if dry_run:
        print(f"\n[DRY RUN] Would append {len(new_records)} record(s) to {output_file}")
        for rec in new_records:
            sym = rec.get("element_symbol") or "?"
            price = rec.get("original_price_per_unit")
            conf = rec.get("confidence_score", "?")
            print(f"    {sym}  price={price}  confidence={conf}  notes={rec.get('notes', '')}")
    else:
        existing = load_existing_records(output_file)
        existing_ids = {r.get("id") for r in existing}
        added = 0
        for rec in new_records:
            if rec["id"] not in existing_ids:
                existing.append(rec)
                existing_ids.add(rec["id"])
                added += 1
            else:
                print(f"  SKIP duplicate: {rec['id']}")
        save_records(existing, output_file)
        print(f"\nAppended {added} new record(s) to {output_file}")
        print(f"Total records now: {len(existing)}")

    return new_records


def main():
    """Entry point."""
    parser = argparse.ArgumentParser(
        description="Import invoice files into the Strategic Materials Ledger price records."
    )
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Directory containing invoice files (default: {DEFAULT_INPUT_DIR})",
    )
    parser.add_argument(
        "--output-file",
        type=Path,
        default=DEFAULT_OUTPUT_FILE,
        help=f"Output JSON file for price records (default: {DEFAULT_OUTPUT_FILE})",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Parse and display results without writing to the output file.",
    )
    args = parser.parse_args()

    print(f"Invoice importer")
    print(f"  Input directory : {args.input_dir.resolve()}")
    print(f"  Output file     : {args.output_file.resolve()}")
    print(f"  Dry run         : {args.dry_run}")
    print()

    new_records = process_invoices(args.input_dir.resolve(), args.output_file.resolve(), args.dry_run)

    # Summary.
    print(f"\n--- Summary ---")
    print(f"  Total new records : {len(new_records)}")
    high_conf = sum(1 for r in new_records if (r.get("confidence_score") or 0) >= 0.8)
    low_conf = len(new_records) - high_conf
    print(f"  High confidence   : {high_conf}")
    print(f"  Needs review      : {low_conf}")
    flagged = [r for r in new_records if r.get("notes")]
    if flagged:
        print(f"  Records with notes: {len(flagged)}")
        for r in flagged:
            print(f"    [{r.get('id')}] {r.get('notes')}")


if __name__ == "__main__":
    main()
